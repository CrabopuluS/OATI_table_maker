"""Минималистичное десктопное приложение для формирования отчёта ОАТИ."""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Dict, List, Optional, Sequence

if __package__ in {None, ""}:
    package_root = Path(__file__).resolve().parent
    parent_directory = package_root.parent
    if str(parent_directory) not in sys.path:
        sys.path.insert(0, str(parent_directory))

from MVP.report_builder import (
    DateRange,
    ReportConfig,
    auto_map_columns,
    build_report,
    build_total_header,
    collect_unique_values,
    describe_data_source,
    extract_unique_dates,
    format_integer,
    format_percent,
    OBJECT_FIELD_DEFINITIONS,
    VIOLATION_FIELD_DEFINITIONS,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - visual hint for the user
    raise SystemExit(
        "Для работы приложения требуется установить пакет openpyxl: pip install openpyxl"
    ) from exc


PLACEHOLDER_LABEL = "— Не выбрано —"
DATA_SOURCE_LABELS = {
    "all": "ОАТИ и ЦАФАП",
    "oati": "Только ОАТИ",
    "cafap": "Только ЦАФАП",
}
LABEL_TO_MODE = {label: mode for mode, label in DATA_SOURCE_LABELS.items()}


class ReportApp(tk.Tk):
    """Главное окно настольного MVP."""

    def __init__(self) -> None:
        super().__init__()
        self.title("Конструктор таблиц ОАТИ — MVP")
        self.geometry("1100x720")
        self.configure(padx=12, pady=12)

        self.violations_path: Optional[str] = None
        self.objects_path: Optional[str] = None
        self.violations_records: List[Dict[str, object]] = []
        self.objects_records: List[Dict[str, object]] = []
        self.violations_columns: List[str] = []
        self.objects_columns: List[str] = []
        self.violation_mapping: Dict[str, str] = {}
        self.object_mapping: Dict[str, str] = {}
        self.report_result = None

        self.violation_mapping_vars: Dict[str, tk.StringVar] = {}
        self.object_mapping_vars: Dict[str, tk.StringVar] = {}

        self.type_mode_var = tk.StringVar(value="all")
        self.violation_mode_var = tk.StringVar(value="all")
        self.data_source_var = tk.StringVar(value="all")
        self.data_source_display_var = tk.StringVar(value=DATA_SOURCE_LABELS["all"])

        self.current_start_var = tk.StringVar()
        self.current_end_var = tk.StringVar()
        self.previous_start_var = tk.StringVar()
        self.previous_end_var = tk.StringVar()

        self.available_dates: List[date] = []
        self.available_types: List[str] = []
        self.available_violations: List[str] = []

        self._build_layout()

    # ------------------------------ UI CONSTRUCTION ---------------------------
    def _build_layout(self) -> None:
        header = ttk.Label(
            self,
            text="Минимальный офлайн-инструмент для формирования отчётной таблицы",
            font=("TkDefaultFont", 14, "bold"),
        )
        header.pack(anchor="w", pady=(0, 8))

        files_frame = ttk.LabelFrame(self, text="Исходные файлы")
        files_frame.pack(fill="x", pady=(0, 10))
        self._build_file_section(files_frame)

        mapping_frame = ttk.Frame(self)
        mapping_frame.pack(fill="x", pady=(0, 10))
        self._build_mapping_section(mapping_frame)

        controls_frame = ttk.LabelFrame(self, text="Параметры отчёта")
        controls_frame.pack(fill="x", pady=(0, 10))
        self._build_controls_section(controls_frame)

        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True)
        self._build_table_section(table_frame)

    def _build_file_section(self, parent: ttk.Frame) -> None:
        violations_button = ttk.Button(
            parent, text="Выбрать выгрузку нарушений", command=self.load_violations
        )
        violations_button.grid(row=0, column=0, padx=6, pady=6, sticky="w")
        self.violations_label = ttk.Label(parent, text="Файл не выбран", width=60)
        self.violations_label.grid(row=0, column=1, sticky="w")

        objects_button = ttk.Button(
            parent, text="Выбрать перечень объектов", command=self.load_objects
        )
        objects_button.grid(row=1, column=0, padx=6, pady=6, sticky="w")
        self.objects_label = ttk.Label(parent, text="Файл не выбран", width=60)
        self.objects_label.grid(row=1, column=1, sticky="w")

    def _build_mapping_section(self, parent: ttk.Frame) -> None:
        violations_frame = ttk.LabelFrame(parent, text="Нарушения")
        violations_frame.pack(side="left", fill="both", expand=True, padx=(0, 8))
        self.violations_mapping_container = ttk.Frame(violations_frame)
        self.violations_mapping_container.pack(fill="both", expand=True, padx=6, pady=6)

        objects_frame = ttk.LabelFrame(parent, text="Объекты")
        objects_frame.pack(side="left", fill="both", expand=True)
        self.objects_mapping_container = ttk.Frame(objects_frame)
        self.objects_mapping_container.pack(fill="both", expand=True, padx=6, pady=6)

    def _build_controls_section(self, parent: ttk.LabelFrame) -> None:
        dates_frame = ttk.Frame(parent)
        dates_frame.pack(fill="x", pady=6)
        self._build_dates_controls(dates_frame)

        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill="x", pady=6)
        self._build_filters(filter_frame)

        action_frame = ttk.Frame(parent)
        action_frame.pack(fill="x", pady=(8, 0))
        build_button = ttk.Button(action_frame, text="Сформировать отчёт", command=self.calculate_report)
        build_button.pack(side="left")
        export_button = ttk.Button(action_frame, text="Сохранить в Excel", command=self.export_report)
        export_button.pack(side="left", padx=6)
        self.status_label = ttk.Label(action_frame, text="")
        self.status_label.pack(side="left", padx=12)

    def _build_dates_controls(self, parent: ttk.Frame) -> None:
        ttk.Label(parent, text="Отчётный период: с").grid(row=0, column=0, sticky="w")
        self.current_start_combo = ttk.Combobox(parent, textvariable=self.current_start_var, width=12, state="readonly")
        self.current_start_combo.grid(row=0, column=1, padx=(4, 10))
        ttk.Label(parent, text="по").grid(row=0, column=2, sticky="w")
        self.current_end_combo = ttk.Combobox(parent, textvariable=self.current_end_var, width=12, state="readonly")
        self.current_end_combo.grid(row=0, column=3, padx=(4, 20))

        ttk.Label(parent, text="Предыдущий период: с").grid(row=0, column=4, sticky="w")
        self.previous_start_combo = ttk.Combobox(parent, textvariable=self.previous_start_var, width=12, state="readonly")
        self.previous_start_combo.grid(row=0, column=5, padx=(4, 10))
        ttk.Label(parent, text="по").grid(row=0, column=6, sticky="w")
        self.previous_end_combo = ttk.Combobox(parent, textvariable=self.previous_end_var, width=12, state="readonly")
        self.previous_end_combo.grid(row=0, column=7, padx=(4, 0))

    def _build_filters(self, parent: ttk.Frame) -> None:
        type_frame = ttk.Frame(parent)
        type_frame.pack(side="left", fill="both", expand=True, padx=(0, 6))
        ttk.Label(type_frame, text="Типы объектов").pack(anchor="w")
        ttk.Radiobutton(
            type_frame,
            text="Все типы",
            variable=self.type_mode_var,
            value="all",
            command=self._on_filter_change,
        ).pack(anchor="w")
        ttk.Radiobutton(
            type_frame,
            text="Выбрать вручную",
            variable=self.type_mode_var,
            value="custom",
            command=self._on_filter_change,
        ).pack(anchor="w")
        self.type_listbox = tk.Listbox(type_frame, selectmode=tk.MULTIPLE, height=6, exportselection=False)
        self.type_listbox.pack(fill="x", pady=(4, 0))

        violation_frame = ttk.Frame(parent)
        violation_frame.pack(side="left", fill="both", expand=True, padx=(0, 6))
        ttk.Label(violation_frame, text="Нарушения").pack(anchor="w")
        ttk.Radiobutton(
            violation_frame,
            text="Все нарушения",
            variable=self.violation_mode_var,
            value="all",
            command=self._on_filter_change,
        ).pack(anchor="w")
        ttk.Radiobutton(
            violation_frame,
            text="Выбрать вручную",
            variable=self.violation_mode_var,
            value="custom",
            command=self._on_filter_change,
        ).pack(anchor="w")
        self.violation_listbox = tk.Listbox(violation_frame, selectmode=tk.MULTIPLE, height=6, exportselection=False)
        self.violation_listbox.pack(fill="x", pady=(4, 0))

        data_source_frame = ttk.Frame(parent)
        data_source_frame.pack(side="left", fill="both", expand=True)
        ttk.Label(data_source_frame, text="Источник данных").pack(anchor="w")
        self.data_source_menu = ttk.Combobox(
            data_source_frame,
            textvariable=self.data_source_display_var,
            state="readonly",
            values=list(DATA_SOURCE_LABELS.values()),
        )
        self.data_source_menu.pack(anchor="w", pady=(4, 0))
        self.data_source_menu.bind("<<ComboboxSelected>>", self._on_data_source_selected)

    def _build_table_section(self, parent: ttk.Frame) -> None:
        columns = (
            "label",
            "total",
            "inspected",
            "inspected_percent",
            "violation_percent",
            "total_violations",
            "current",
            "previous",
            "resolved",
            "on_control",
        )
        self.tree = ttk.Treeview(parent, columns=columns, show="headings", height=12)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        for column in columns:
            self.tree.heading(column, text="")
            self.tree.column(column, anchor="center", width=100)
        self.tree.column("label", anchor="w", width=180)

    # ------------------------------ FILE LOADING ------------------------------
    def load_violations(self) -> None:
        path = filedialog.askopenfilename(
            title="Выберите файл с нарушениями",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")],
        )
        if not path:
            return
        try:
            columns, records = read_excel_records(path)
        except Exception as exc:  # pylint: disable=broad-except
            messagebox.showerror("Ошибка", f"Не удалось прочитать файл нарушений: {exc}")
            return
        self.violations_path = path
        self.violations_columns = columns
        self.violations_records = records
        self.violation_mapping = auto_map_columns(columns, VIOLATION_FIELD_DEFINITIONS)
        self.violations_label.configure(text=path)
        self._render_mapping_controls(
            self.violations_mapping_container,
            VIOLATION_FIELD_DEFINITIONS,
            columns,
            self.violation_mapping,
            self.violation_mapping_vars,
        )
        self._update_filters()
        self.status_label.configure(text="")

    def load_objects(self) -> None:
        path = filedialog.askopenfilename(
            title="Выберите файл с перечнем объектов",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")],
        )
        if not path:
            return
        try:
            columns, records = read_excel_records(path)
        except Exception as exc:  # pylint: disable=broad-except
            messagebox.showerror("Ошибка", f"Не удалось прочитать перечень объектов: {exc}")
            return
        self.objects_path = path
        self.objects_columns = columns
        self.objects_records = records
        self.object_mapping = auto_map_columns(columns, OBJECT_FIELD_DEFINITIONS)
        self.objects_label.configure(text=path)
        self._render_mapping_controls(
            self.objects_mapping_container,
            OBJECT_FIELD_DEFINITIONS,
            columns,
            self.object_mapping,
            self.object_mapping_vars,
        )
        self._update_filters()
        self.status_label.configure(text="")

    def _render_mapping_controls(
        self,
        container: ttk.Frame,
        definitions: Sequence,
        columns: Sequence[str],
        mapping: Dict[str, str],
        variables: Dict[str, tk.StringVar],
    ) -> None:
        for widget in container.winfo_children():
            widget.destroy()
        variables.clear()
        options = [PLACEHOLDER_LABEL] + list(columns)
        for definition in definitions:
            row = ttk.Frame(container)
            row.pack(fill="x", pady=2)
            label = ttk.Label(row, text=definition.label, width=38)
            label.pack(side="left")
            var = tk.StringVar()
            current_value = mapping.get(definition.key) or PLACEHOLDER_LABEL
            var.set(current_value)
            combo = ttk.Combobox(row, textvariable=var, values=options, state="readonly", width=30)
            combo.pack(side="left", padx=(6, 0))

            def make_handler(field_key: str, variable: tk.StringVar) -> None:
                def _handler(*_args: str) -> None:
                    value = variable.get()
                    mapping[field_key] = "" if value == PLACEHOLDER_LABEL else value
                    self._update_filters()
                variable.trace_add("write", _handler)

            make_handler(definition.key, var)
            variables[definition.key] = var

    # ------------------------------ FILTERS -----------------------------------
    def _on_filter_change(self) -> None:
        self.status_label.configure(text="")

    def _on_data_source_selected(self, _event: tk.Event) -> None:  # type: ignore[override]
        label = self.data_source_display_var.get()
        self.data_source_var.set(LABEL_TO_MODE.get(label, "all"))
        self._on_filter_change()

    def _update_filters(self) -> None:
        if not self.violations_records:
            return
        date_column = self.violation_mapping.get("inspectionDate")
        self.available_dates = extract_unique_dates(self.violations_records, date_column)
        formatted_dates = [format_date_display(value) for value in self.available_dates]
        for combo, var in (
            (self.current_start_combo, self.current_start_var),
            (self.current_end_combo, self.current_end_var),
            (self.previous_start_combo, self.previous_start_var),
            (self.previous_end_combo, self.previous_end_var),
        ):
            combo.configure(values=formatted_dates)
            if formatted_dates and not var.get():
                var.set(formatted_dates[0])
        if formatted_dates:
            self.current_end_var.set(self.current_end_var.get() or formatted_dates[-1])
            self.previous_end_var.set(self.previous_end_var.get() or formatted_dates[-1])

        type_column = self.violation_mapping.get("objectType")
        self.available_types = collect_unique_values(self.violations_records, type_column)
        self._refresh_listbox(self.type_listbox, self.available_types)

        violation_column = self.violation_mapping.get("violationName")
        self.available_violations = collect_unique_values(
            self.violations_records, violation_column
        )
        self._refresh_listbox(self.violation_listbox, self.available_violations)

    def _refresh_listbox(self, listbox: tk.Listbox, values: Sequence[str]) -> None:
        current_selection = listbox.curselection()
        listbox.delete(0, tk.END)
        for value in values:
            listbox.insert(tk.END, value)
        for index in current_selection:
            if index < len(values):
                listbox.selection_set(index)

    # ------------------------------ REPORT ------------------------------------
    def calculate_report(self) -> None:
        if not self.violations_records or not self.objects_records:
            messagebox.showinfo("Требуется ввод", "Загрузите оба исходных файла перед расчётом")
            return
        try:
            config = self._build_config()
        except ValueError as exc:
            messagebox.showerror("Ошибка", str(exc))
            return
        try:
            self.report_result = build_report(
                self.violations_records, self.objects_records, config
            )
        except ValueError as exc:
            messagebox.showerror("Ошибка", str(exc))
            return
        self._render_report_table(config)
        self.status_label.configure(text="Отчёт сформирован")

    def _build_config(self) -> ReportConfig:
        for definition in VIOLATION_FIELD_DEFINITIONS:
            if not definition.optional and not self.violation_mapping.get(definition.key):
                raise ValueError(
                    f"Не заполнено сопоставление поля: {definition.label}"
                )
        for definition in OBJECT_FIELD_DEFINITIONS:
            if not definition.optional and not self.object_mapping.get(definition.key):
                raise ValueError(f"Не заполнено сопоставление поля: {definition.label}")

        current_period = self._build_period(self.current_start_var.get(), self.current_end_var.get(), "отчётного")
        previous_period = self._build_period(
            self.previous_start_var.get(), self.previous_end_var.get(), "предыдущего"
        )

        if current_period.start > current_period.end:
            raise ValueError("Дата начала отчётного периода позже даты окончания")
        if previous_period.start > previous_period.end:
            raise ValueError("Дата начала предыдущего периода позже даты окончания")

        selected_types = self._extract_selection(self.type_listbox, self.available_types)
        if self.type_mode_var.get() == "custom" and len(selected_types) > 3:
            raise ValueError("Можно выбрать не более трёх типов объектов")
        selected_violations = self._extract_selection(
            self.violation_listbox, self.available_violations
        )

        return ReportConfig(
            violation_mapping=self.violation_mapping.copy(),
            object_mapping=self.object_mapping.copy(),
            current_period=current_period,
            previous_period=previous_period,
            type_mode=self.type_mode_var.get(),
            selected_types=selected_types,
            violation_mode=self.violation_mode_var.get(),
            selected_violations=selected_violations,
            data_source_mode=self.data_source_var.get(),
        )

    def _build_period(self, start_text: str, end_text: str, label: str) -> DateRange:
        if not start_text or not end_text:
            raise ValueError(f"Укажите границы {label} периода")
        start_date = parse_date_display(start_text)
        end_date = parse_date_display(end_text)
        if not start_date or not end_date:
            raise ValueError(f"Некорректные даты {label} периода")
        return DateRange(start=start_date, end=end_date)

    def _extract_selection(
        self, listbox: tk.Listbox, values: Sequence[str]
    ) -> List[str]:
        return [values[index] for index in listbox.curselection() if index < len(values)]

    def _render_report_table(self, config: ReportConfig) -> None:
        for row in self.tree.get_children():
            self.tree.delete(row)
        headers = build_table_headers(config)
        for column_id, header_text in zip(self.tree.cget("columns"), headers):
            self.tree.heading(column_id, text=header_text)
        if not self.report_result:
            return
        for row in self.report_result.rows:
            self.tree.insert("", tk.END, values=self._format_row(row))
        self.tree.insert("", tk.END, values=self._format_row(self.report_result.total_row))

    def _format_row(self, row) -> Sequence[str]:
        return (
            row.label,
            format_integer(row.total_objects),
            format_integer(row.inspected_objects),
            format_percent(row.inspected_percent),
            format_percent(row.violation_percent),
            format_integer(row.total_violations),
            format_integer(row.current_violations),
            format_integer(row.previous_control),
            format_integer(row.resolved),
            format_integer(row.on_control),
        )

    # ------------------------------ EXPORT ------------------------------------
    def export_report(self) -> None:
        if not self.report_result:
            messagebox.showinfo("Отчёт отсутствует", "Сначала сформируйте отчёт")
            return
        try:
            config = self._build_config()
        except ValueError as exc:
            messagebox.showerror("Ошибка", str(exc))
            return
        path = filedialog.asksaveasfilename(
            title="Сохранить отчёт",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            export_report_to_excel(path, self.report_result, config)
        except Exception as exc:  # pylint: disable=broad-except
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {exc}")
            return
        messagebox.showinfo("Готово", "Отчёт сохранён")


# ------------------------------ HELPER FUNCTIONS -----------------------------

def read_excel_records(path: str) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        return [], []
    headers = [get_cell_text(cell) for cell in rows[0]]
    normalized_headers = [header or "" for header in headers]
    records: List[Dict[str, object]] = []
    for row in rows[1:]:
        record: Dict[str, object] = {}
        has_value = False
        for header, cell in zip(normalized_headers, row):
            value = extract_cell_value(cell)
            record[header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            records.append(record)
    return normalized_headers, records


def get_cell_text(cell) -> str:
    value = extract_cell_value(cell)
    if value in (None, ""):
        return ""
    return str(value)


def extract_cell_value(cell) -> object:
    if cell is None:
        return ""
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def parse_date_display(text: str) -> Optional[date]:
    try:
        return datetime.strptime(text, "%d.%m.%Y").date()
    except ValueError:
        return None


def format_date_display(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def build_table_headers(config: ReportConfig) -> Sequence[str]:
    total_header = build_total_header(config.type_mode, config.selected_types)
    range_header = (
        f"Проверено ОДХ с {format_date_display(config.current_period.start)} "
        f"по {format_date_display(config.current_period.end)}"
    )
    return (
        "Округ",
        total_header,
        range_header,
        "% проверенных объектов от общего количества ОДХ",
        "% объектов с нарушениями",
        "Всего нарушений",
        "Нарушения, выявленные за отчётный период",
        "Нарушения, на контроле с предыдущей проверки",
        "Устранено нарушений",
        "Нарушения на контроле",
    )


def export_report_to_excel(path: str, report, config: ReportConfig) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"

    title = build_excel_title(config)
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    sheet.append([])
    sheet.append(build_excel_header_numbers(10))
    headers = build_table_headers(config)
    sheet.append(headers)

    for row in report.rows:
        sheet.append(_export_row_values(row))
    sheet.append(_export_row_values(report.total_row))

    for row in sheet.iter_rows(min_row=5, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "0.0%"

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    workbook.save(path)


def build_excel_title(config: ReportConfig) -> str:
    period = config.current_period
    period_text = f"{format_date_display(period.start)}–{format_date_display(period.end)}"
    parts = [f"Нарушения на ОДХ (отчёт за {period_text})"]
    data_source = describe_data_source(config.data_source_mode)
    if data_source:
        parts.append(data_source)
    return ", ".join(parts)


def build_excel_header_numbers(column_count: int) -> List[str]:
    template = ["1", "2", "3", "4", "4.1", "4.2", "5", "6", "6.1", "6.2"]
    if column_count <= len(template):
        return template[:column_count]
    result = template[:]
    for index in range(len(template), column_count):
        result.append(str(index + 1))
    return result


def _export_row_values(row) -> List[object]:
    return [
        row.label,
        row.total_objects,
        row.inspected_objects,
        row.inspected_percent / 100 if row.inspected_percent else 0,
        row.violation_percent / 100 if row.violation_percent else 0,
        row.total_violations,
        row.current_violations,
        row.previous_control,
        row.resolved,
        row.on_control,
    ]


def main() -> None:
    app = ReportApp()
    app.mainloop()


if __name__ == "__main__":
    main()
